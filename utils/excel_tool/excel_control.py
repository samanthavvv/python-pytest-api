"""
#!/usr/bin/env python
# -*- coding:utf-8 -*-

@Author : Ada
@Time : 2023/4/27 14:53
@Project : ada_pytest_api_project
@File : excel_control.py

"""
# Code starts here...
import json
import os
from openpyxl import load_workbook

from utils.format_data_tool.models import TestCaseEnum, TestCase



class GetExcelData:
    """
    从excel中读取数据，并校验和输出
    """

    def __init__(self, file_path):
        self.file_path = file_path

    def format_data(self, values:str):
        if isinstance(values,str):
            new_value = values.replace('\n','').replace('\t','')
            try:
                new_value = json.loads(new_value)
            except:
                pass
            return new_value
        return values

    def orgnize_data(self):
        # [{'sheet_name':xx, 'feature':xx, 'case_id1':xx, 'case_id2':xx,...}]
        sheet_list = []

        if os.path.exists(self.file_path):
            wb = load_workbook(self.file_path)

            for name in wb.sheetnames:
                dict_data = {}
                sheet = wb[name]
                # 获取每个sheet 最大行数
                max_row_num = sheet.max_row
                # 获取每个sheet 最大列数
                max_col_num = len(TestCaseEnum)
                key_list = [key.value[0] for key in TestCaseEnum]

                dict_data['sheet_name'] = name
                dict_data['feature'] = sheet.cell(column=1, row=2).value    # featur 和sheet_name 都是以一个sheet 为单位的

                # 从第2行开始，遍历每一行
                for row in range(2, max_row_num+1):
                    case_id = sheet.cell(column=4, row=row).value
                    dict_data[case_id] = {}
                    for col in range(5, max_col_num+1):
                        k = key_list[col-1]
                        v = sheet.cell(column=col, row=row).value
                        dict_data[case_id][k] = self.format_data(v)
                sheet_list.append(dict_data)

            return sheet_list

if __name__ == '__main__':
    r = r'E:\allProject\ada_pytest_api_project\data\YunKuaiJi\DebugTest\debug_test_query.xlsx'
    re = GetExcelData(r).orgnize_data()
    print(re)
